"""
CLI entry point for the SOV Excel parser.

Configuration defaults to ``config/config.yaml``. Sample workbooks live under
``samples/``.

Usage:
    python main.py <file_or_folder>
    python main.py samples/sample_sov.xlsx
    python main.py sov_file.xlsx --profile-arn arn:aws:bedrock:...
    python main.py --rebuild-embeddings
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
import warnings
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

from sov.engine import (
    configure_logging,
    load_config,
    parse_sov_file,
    rebuild_embedding_cache,
)

warnings.filterwarnings("ignore")

SCRIPT_DIR = Path(__file__).resolve().parent
logger = logging.getLogger(__name__)


def _resolve(path_str: str) -> str:
    """Return an absolute path, resolving relative paths against ``SCRIPT_DIR``."""
    p = Path(path_str)
    if not p.is_absolute():
        p = SCRIPT_DIR / p
    return str(p)


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Define and parse CLI arguments; ``argv`` defaults to ``sys.argv`` when omitted."""
    p = argparse.ArgumentParser(description="Parse property SOV Excel file(s).")
    p.add_argument("path", nargs="?", default=".", help=".xlsx/.xls file OR folder")
    p.add_argument(
        "--config",
        default=None,
        help="Path to main YAML config (default: config/config.yaml)",
    )
    p.add_argument("--profile-arn", dest="inference_profile_arn", help="Bedrock inference profile ARN")
    p.add_argument("--region", dest="aws_region", help="AWS region")
    p.add_argument("--aws-profile", dest="aws_profile", help="Named AWS CLI profile")
    p.add_argument("--rebuild-embeddings", action="store_true", help="Force rebuild embedding cache")
    p.add_argument("--sheets", default=None, help="Comma-separated sheet indices or names")
    p.add_argument(
        "--verbose",
        "-v",
        action="store_true",
        help="Log detailed SOV engine diagnostics (default: WARNING+ for engine, INFO for CLI)",
    )
    return p.parse_args(argv)


def _collect_files(path: str) -> list[Path]:
    """Collect ``.xlsx`` / ``.xls`` files from a file path or directory (non-recursive)."""
    p = Path(path)
    if p.is_file():
        return [p]
    if p.is_dir():
        return sorted(
            f
            for f in p.iterdir()
            if f.suffix.lower() in (".xlsx", ".xls") and not f.name.startswith("~$")
        )
    raise FileNotFoundError(f"Path not found: {path}")


def _append_sheets_to_source(
    source_path: str,
    df: pd.DataFrame,
    df_sources: pd.DataFrame,
) -> None:
    """Append ``Cleaned Data`` and ``Source References`` sheets to an existing workbook."""
    wb = load_workbook(source_path)
    for name in ("Cleaned Data", "Source References"):
        if name in wb.sheetnames:
            del wb[name]
    wb.save(source_path)

    with pd.ExcelWriter(source_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        # Convert all NA variants to empty string for clean Excel output
        df_out = df.copy()
        for col in df_out.columns:
            df_out[col] = df_out[col].astype(object).where(df_out[col].notna(), "")
            df_out[col] = df_out[col].replace({"<NA>": "", "nan": "", "None": "", "NaN": ""})
        df_src_out = df_sources.copy()
        for col in df_src_out.columns:
            df_src_out[col] = df_src_out[col].astype(object).where(df_src_out[col].notna(), "")
            df_src_out[col] = df_src_out[col].replace({"<NA>": "", "nan": "", "None": "", "NaN": ""})
        df_out.to_excel(writer, sheet_name="Cleaned Data", index=False)
        df_src_out.to_excel(writer, sheet_name="Source References", index=False)


def process_file(
    filepath: str,
    config_path: str | None,
    overrides: dict | None,
    output_dir: str | None = None,
) -> tuple[bool, str]:
    """Run the SOV pipeline on one workbook and write JSON metadata.

    Parameters
    ----------
    filepath
        Path to ``.xlsx`` or ``.xls``.
    config_path
        Optional path to YAML config; ``None`` uses engine default.
    overrides
        Bedrock / sheet overrides merged into config.
    output_dir
        Directory for ``*_llm_responses.json``; if ``None``, JSON is next to the input file.

    Returns
    -------
    tuple[bool, str]
        ``(True, "")`` on success; ``(False, error_message)`` on failure.
    """
    try:
        df, df_sources, metadata = parse_sov_file(
            filepath,
            config_path=config_path,
            config_overrides=overrides,
            verbose=False,
        )

        if df.empty:
            return False, "No data extracted"

        _append_sheets_to_source(filepath, df, df_sources)

        # Parse raw_response strings into dicts for readable JSON output
        for resp in metadata.get("llm_responses", []):
            raw = resp.get("raw_response", "")
            if isinstance(raw, str):
                try:
                    resp["raw_response"] = json.loads(raw)
                except (json.JSONDecodeError, TypeError):
                    pass  # keep as string if not valid JSON

        fname = Path(filepath).stem + "_llm_responses.json"
        out_json = (
            str(Path(output_dir) / fname)
            if output_dir
            else filepath.rsplit(".", 1)[0] + "_llm_responses.json"
        )
        Path(out_json).parent.mkdir(parents=True, exist_ok=True)
        with open(out_json, "w", encoding="utf-8") as f:
            json.dump(metadata, f, indent=2, ensure_ascii=False, default=str)

        return True, ""

    except Exception as exc:  # pylint: disable=broad-exception-caught
        # Pipeline may fail on I/O, AWS, or parsing; return message for batch summary.
        return False, str(exc)


def main(argv: list[str] | None = None) -> None:
    """Parse CLI arguments, optionally rebuild embeddings, or process SOV files with a progress bar.

    Exits with code 1 if no spreadsheets are found under the given path.
    """
    args = _parse_args(argv)

    # Console logging: always INFO for this CLI; quiet mode suppresses ``sov`` engine detail
    configure_logging(level=logging.INFO)
    if not args.verbose:
        logging.getLogger("sov").setLevel(logging.WARNING)

    target = _resolve(args.path)
    config_path = _resolve(args.config) if args.config else None

    overrides: dict = {}
    if args.inference_profile_arn:
        overrides["inference_profile_arn"] = args.inference_profile_arn
    if args.aws_region:
        overrides["aws_region"] = args.aws_region
    if args.aws_profile:
        overrides["aws_profile"] = args.aws_profile
    if args.sheets:
        parts = [s.strip() for s in args.sheets.split(",")]
        overrides["sheets"] = [int(s) if s.isdigit() else s for s in parts]

    # Rebuild embeddings if explicitly asked, then exit
    if args.rebuild_embeddings:
        logging.getLogger("sov").setLevel(logging.INFO)
        cfg = load_config(config_path)
        rebuild_embedding_cache(cfg)
        logger.info("Embedding cache rebuilt.")
        return

    files = _collect_files(target)
    if not files:
        logger.error("No .xlsx/.xls files found in %s", target)
        sys.exit(1)

    # Output directory
    input_dir = Path(target) if Path(target).is_dir() else Path(target).parent
    output_dir = str(input_dir / "output")
    Path(output_dir).mkdir(exist_ok=True)

    # Process with tqdm progress bar
    success = 0
    failed = 0
    errors: list[tuple[str, str]] = []

    pbar = tqdm(files, desc="Processing SOV files", unit="file", ncols=100)
    for f in pbar:
        pbar.set_postfix_str(f.name[:30], refresh=True)
        ok, err = process_file(str(f), config_path, overrides or None, output_dir=output_dir)
        if ok:
            success += 1
        else:
            failed += 1
            errors.append((f.name, err))
        pbar.set_postfix(ok=success, fail=failed)

    pbar.close()

    logger.info("%s", "=" * 60)
    logger.info(
        "Done: %d succeeded, %d failed, %d total",
        success,
        failed,
        len(files),
    )
    logger.info("JSON output directory: %s/", output_dir)
    logger.info("Cleaned Data / Source References sheets appended to each source workbook")
    logger.info("%s", "=" * 60)

    if errors:
        logger.error("Failures (%d):", len(errors))
        for fname, err in errors:
            logger.error("  %s: %s", fname, err)


if __name__ == "__main__":
    main()
