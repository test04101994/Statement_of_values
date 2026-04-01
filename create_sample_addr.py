"""Create a sample SOV file with only address column — no separate town/state/zip/country."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SOV 2025"

# Header row
headers = [
    "Loc #", "Insured Name", "Address", "Occupancy Type",
    "Construction Type", "Year Built", "Building Value",
    "Contents Value", "BI Value", "Currency"
]
ws.append(headers)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Data rows — international addresses, NO separate town/state/zip/country columns
data = [
    [1, "GlobalTech Inc", "Ambrosetti 2431, Munro, Buenos Aires, Argentina",
     "Office", "Steel Frame", 2005, 48000000, 13000000, 8500000, "ARS"],
    [2, "Acme Corp", "500 Industrial Blvd, Chicago, IL 60601, USA",
     "Warehouse", "Pre-Eng Metal", 2010, 20000000, 5500000, 3200000, "USD"],
    [3, "TechnoLabs", "45 High Street, London SW1A 1AA, United Kingdom",
     "Laboratory", "Reinforced Concrete", 2018, 35000000, 18000000, 12000000, "GBP"],
    [4, "Bavaria Motors", "Gewerbering 2, 2440 Moosbrunn, Austria",
     "Manufacturing", "Masonry", 1995, 1200000, 800000, 400000, "EUR"],
    [5, "BelgoChem", "Krijgsbaan 247/D1, 9140 Temse, Belgium",
     "Data Center", "Steel Frame", 2020, 55000000, 40000000, 25000000, "EUR"],
    [6, "SantaCruz Mining", "Parque Industrial Manzano #3, Santa Cruz de la Sierra, Bolivia",
     "Manufacturing", "Pre-Eng Metal", 2001, 22000000, 15000000, 8000000, "BOB"],
    [7, "Brasil Logistica", "Rua Barra Longa 11, Betim, Minas Gerais, Brazil",
     "Warehouse", "Masonry", 2012, 9000000, 3000000, 2000000, "BRL"],
    [8, "Maple Industries", "1200 Innovation Dr, Toronto, ON M5V 3L9, Canada",
     "Office", "Steel Frame", 2015, 16000000, 8000000, 6000000, "CAD"],
]

for row in data:
    ws.append(row)

# Totals row
ws.append(["", "TOTAL", "", "", "", "", 206200000, 103300000, 65100000, ""])

out = "sample_sov_addr_only.xlsx"
wb.save(out)
print(f"Created {out} with {len(data)} rows, address-only (no town/state/zip/country columns)")
