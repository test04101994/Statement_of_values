"""
Generate a realistic sample SOV Excel file with multiple sheets.

Run with ``python create_sample.py`` from the project directory. Output is
``sample_sov.xlsx`` in the current working directory.
"""

from __future__ import annotations

import logging

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

OUTPUT_FILENAME = "sample_sov.xlsx"


def _build_workbook() -> openpyxl.Workbook:
    """Create and populate the workbook with cover, SOV, and summary sheets."""
    wb = openpyxl.Workbook()

    # Sheet 1: Cover page (should be skipped by LLM)
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover["A1"] = "ACME Insurance Corp"
    ws_cover["A1"].font = Font(size=18, bold=True)
    ws_cover["A3"] = "Statement of Values — Property Portfolio"
    ws_cover["A5"] = "Prepared: March 2026"
    ws_cover["A6"] = "Broker: Marsh McLennan"
    ws_cover["A7"] = "Insured: GlobalTech Industries Inc."

    # Sheet 2: 2024 SOV (older year — should NOT be picked)
    ws_2024 = wb.create_sheet("SOV 2024")
    headers_2024 = [
        "Loc #",
        "Location Name",
        "Street Address",
        "City",
        "State",
        "ZIP",
        "Occupancy",
        "Construction",
        "Year Built",
        "Stories",
        "Sq Ft",
        "Building Value 2024",
        "Contents Value 2024",
        "BI Value 2024",
        "TIV 2024",
        "Sprinklered",
        "Flood Zone",
    ]
    ws_2024.append(headers_2024)
    data_2024 = [
        [
            1,
            "HQ Office Tower",
            "100 Main St",
            "New York",
            "NY",
            "10001",
            "Office",
            "Steel Frame",
            2005,
            25,
            500000,
            45000000,
            12000000,
            8000000,
            65000000,
            "Yes",
            "X",
        ],
        [
            2,
            "Distribution Center",
            "500 Industrial Blvd",
            "Chicago",
            "IL",
            "60601",
            "Warehouse",
            "Pre-Eng Metal",
            2010,
            2,
            250000,
            18000000,
            5000000,
            3000000,
            26000000,
            "Yes",
            "A",
        ],
    ]
    for row in data_2024:
        ws_2024.append(row)

    # Sheet 3: 2025 SOV (latest year — SHOULD be picked)
    ws_2025 = wb.create_sheet("SOV 2025")
    ws_2025.append(["GlobalTech Industries — 2025 Property Statement of Values"])
    ws_2025.append([])

    headers_2025 = [
        "Loc #",
        "Location Name",
        "Street Address",
        "City",
        "State",
        "ZIP",
        "Country",
        "Occupancy Type",
        "Construction Type",
        "Year Built",
        "# Stories",
        "Square Footage",
        "Building Value 2025",
        "Contents Value 2025",
        "BI Value 2025",
        "Other Value 2025",
        "TIV 2025",
        "Flood Zone",
        "EQ Zone",
        "Sprinklered",
        "Policy Limit",
        "Deductible",
        "Latitude",
        "Longitude",
    ]
    ws_2025.append(headers_2025)

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for cell in ws_2025[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    data_2025 = [
        [
            1,
            "HQ Office Tower",
            "100 Main St",
            "New York",
            "NY",
            "10001",
            "USA",
            "Office",
            "Steel Frame",
            2005,
            25,
            500000,
            48000000,
            13000000,
            8500000,
            500000,
            70000000,
            "X",
            "3",
            "Yes",
            70000000,
            50000,
            40.7128,
            -74.0060,
        ],
        [
            2,
            "Distribution Center",
            "500 Industrial Blvd",
            "Chicago",
            "IL",
            "60601",
            "USA",
            "Warehouse",
            "Pre-Eng Metal",
            2010,
            2,
            "250,000",
            "$20,000,000",
            "$5,500,000",
            "$3,200,000",
            "$0",
            "$28,700,000",
            "A",
            "1",
            "Y",
            28700000,
            25000,
            41.8781,
            -87.6298,
        ],
        [
            3,
            "R&D Campus Bldg A",
            "1200 Innovation Dr",
            "San Jose",
            "CA",
            "95110",
            "USA",
            "Laboratory",
            "Reinforced Concrete",
            2018,
            4,
            120000,
            35000000,
            18000000,
            12000000,
            2000000,
            67000000,
            "X",
            "5",
            "Partial",
            67000000,
            100000,
            37.3382,
            -121.8863,
        ],
        [
            4,
            "Retail Storefront",
            "88 Commerce Ave",
            "Dallas",
            "TX",
            "75201",
            "USA",
            "Retail",
            "Masonry",
            "circa 1995",
            "one",
            8000,
            "TBD",
            800000,
            400000,
            "N/A",
            2400000,
            "C",
            "0",
            "No",
            2400000,
            10000,
            32.7767,
            999.0,
        ],
        [
            5,
            "Data Center",
            "2000 Server Rd",
            "Ashburn",
            "VA",
            "20147",
            "USA",
            "Data Center",
            "Steel Frame",
            2020,
            3,
            60000,
            55000000,
            40000000,
            25000000,
            "(5000)",
            125000000,
            "X",
            "2",
            "true",
            100000000,
            250000,
            39.0438,
            -77.4874,
        ],
        [
            6,
            "Manufacturing Plant",
            "750 Factory Ln",
            "Detroit",
            "MI",
            "48201",
            "USA",
            "Manufacturing",
            "Pre-Eng Metal",
            2001,
            2,
            180000,
            22000000,
            15000000,
            8000000,
            1000000,
            46000000,
            "B",
            "1",
            "1",
            46000000,
            75000,
            42.3314,
            -83.0458,
        ],
        [
            7,
            "Regional Office",
            "300 Park Ave",
            "Atlanta",
            "GA",
            "30301",
            "USA",
            "Office",
            "Masonry",
            2012,
            8,
            45000,
            9000000,
            3000000,
            "-",
            None,
            14000000,
            "X",
            "0",
            "n",
            14000000,
            25000,
            33.7490,
            -84.3880,
        ],
        [
            8,
            "Cold Storage Facility",
            "1500 Frost Way",
            "Minneapolis",
            "MN",
            "55401",
            "USA",
            "Cold Storage",
            "Insulated Metal Panel",
            2015,
            1,
            95000,
            16000000,
            8000000,
            6000000,
            500000,
            30500000,
            "A",
            "0",
            "Yes",
            30500000,
            50000,
            44.9778,
            -93.2650,
        ],
    ]

    for row in data_2025:
        ws_2025.append(row)

    for r in range(4, 12):
        for c in [13, 14, 15, 16, 17, 21, 22]:
            cell = ws_2025.cell(row=r, column=c)
            cell.number_format = "$#,##0"

    ws_2025.append([])
    ws_2025.append(
        [
            "",
            "TOTAL",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            206200000,
            103300000,
            65100000,
            9000000,
            383600000,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )

    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Category", "Count", "Total TIV"])
    ws_summary.append(["Office", 2, 84000000])
    ws_summary.append(["Warehouse", 1, 28700000])
    ws_summary.append(["Manufacturing", 1, 46000000])
    ws_summary.append(["Data Center", 1, 125000000])
    ws_summary.append(["Other", 3, 99900000])

    return wb


def main() -> None:
    """Write ``OUTPUT_FILENAME`` and log sheet names."""
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    wb = _build_workbook()
    wb.save(OUTPUT_FILENAME)
    logger.info("Created %s with sheets: %s", OUTPUT_FILENAME, wb.sheetnames)


if __name__ == "__main__":
    main()
