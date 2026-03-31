"""
SOV (Statement of Values) parsing engine.

Single module that handles: config loading, Excel I/O, Bedrock LLM calls,
sheet auto-detection, DataFrame building, column derivations, and multi-year
validation.  Add new output columns by editing ``config.yaml`` — no code
changes required.
"""

from __future__ import annotations

import json
import logging
import os
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import boto3
import openpyxl
import pandas as pd
import xlrd
import yaml
from botocore.config import Config as BotoConfig

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

_LOG_FMT = "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s"
logger = logging.getLogger("sov")


def configure_logging(level: int = logging.INFO) -> None:
    """Configure the root logger with a stderr handler if none exist.

    Parameters
    ----------
    level
        Minimum log level for the root logger (default: :data:`logging.INFO`).
    """
    root = logging.getLogger()
    if not root.handlers:
        handler = logging.StreamHandler(sys.stderr)
        handler.setFormatter(logging.Formatter(_LOG_FMT))
        root.addHandler(handler)
    root.setLevel(level)


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

_SCRIPT_DIR = Path(__file__).resolve().parent.parent
_DEFAULT_CONFIG = _SCRIPT_DIR / "config.yaml"
_DEFAULT_REGION = "us-east-1"
_DEFAULT_MODEL = "us.anthropic.claude-sonnet-4-20250514-v1:0"


def load_config(path: Union[str, Path, None] = None) -> Dict[str, Any]:
    """Load unified SOV configuration from YAML.

    Parses ``fields`` into descriptions, optional ``type`` and ``synonyms`` per field,
    merges ``learned_synonyms.yaml`` when present, and returns augmented keys
    ``field_types`` and ``synonyms`` alongside ``fields``.

    Parameters
    ----------
    path
        Path to ``config.yaml``. Defaults to ``config.yaml`` next to the project root.

    Returns
    -------
    dict
        Configuration including ``fields``, ``field_types``, ``synonyms``, ``bedrock``, etc.

    Raises
    ------
    FileNotFoundError
        If the config file does not exist.
    ValueError
        If ``fields`` is missing or empty.
    """
    p = Path(path) if path else _DEFAULT_CONFIG
    if not p.is_file():
        raise FileNotFoundError(f"Config not found: {p}")
    with open(p, encoding="utf-8") as f:
        cfg = yaml.safe_load(f) or {}

    # Parse fields — supports both formats:
    #   field_name: "description"                  (flat, type defaults to text)
    #   field_name: {desc: "description", type: "amount"}
    raw_fields = cfg.get("fields")
    if not isinstance(raw_fields, dict) or not raw_fields:
        raise ValueError(f"'fields' must be a non-empty mapping in {p}")

    schema: dict[str, str] = {}  # field → description (for LLM)
    field_types: dict[str, str] = {}  # field → type (for validation)
    synonyms: dict[str, list[str]] = {}  # field → list of known header names

    for k, v in raw_fields.items():
        if isinstance(v, dict):
            desc = v.get("desc", "")
            ftype = v.get("type", "text")
            syns = v.get("synonyms", [])
        else:
            desc = str(v) if v else ""
            ftype = "text"
            syns = []
        if not desc.strip():
            raise ValueError(f"Field '{k}' must have a non-empty description in {p}")
        schema[k] = desc.strip()
        field_types[k] = ftype.strip().lower()
        synonyms[k] = [str(s).strip() for s in syns if s]

    # Load learned synonyms (human feedback file)
    learned_path = p.parent / "learned_synonyms.yaml"
    if learned_path.is_file():
        with open(learned_path, encoding="utf-8") as f:
            learned = yaml.safe_load(f) or {}
        for field, syns_list in learned.items():
            if field in synonyms and isinstance(syns_list, list):
                existing = {s.lower() for s in synonyms[field]}
                for s in syns_list:
                    if isinstance(s, dict):
                        s = s.get("name", "")
                    s = str(s).strip()
                    if s and s.lower() not in ("none", "nan", "") and s.lower() not in existing:
                        synonyms[field].append(s)
                        existing.add(s.lower())
        logger.info("Loaded learned synonyms from %s", learned_path)

    cfg["fields"] = schema
    cfg["field_types"] = field_types
    cfg["synonyms"] = synonyms

    return cfg


# ---------------------------------------------------------------------------
# Excel I/O
# ---------------------------------------------------------------------------


_CURRENCY_RE = re.compile(r"[$ ,€£¥₹%]")


def _col_letter(col_idx: int) -> str:
    """Convert 0-based column index to Excel column letter (0→A, 25→Z, 26→AA)."""
    result = ""
    idx = col_idx
    while True:
        result = chr(65 + idx % 26) + result
        idx = idx // 26 - 1
        if idx < 0:
            break
    return result


def _extract_cell_value(cell) -> object:
    """Extract cell value, recovering numeric values hidden behind formatting.

    openpyxl ``data_only=True`` usually returns the raw number, but some files
    store formatted strings like ``"$1,200,000"`` or ``"15%"``.  This strips
    currency/grouping symbols and returns a float when the underlying value is
    numeric.
    """
    val = cell.value
    if val is None:
        return None
    # Already numeric — keep as-is
    if isinstance(val, (int, float)):
        return val
    # String with currency/formatting artifacts → recover numeric value.
    # Only convert if the original string contains formatting characters
    # ($, commas, €, etc.) — bare strings like "1" or "Yes" stay as strings.
    if isinstance(val, str):
        raw = val.strip()
        if _CURRENCY_RE.search(raw) or (raw.startswith("(") and raw.endswith(")")):
            stripped = _CURRENCY_RE.sub("", raw)
            if stripped.startswith("(") and stripped.endswith(")"):
                stripped = "-" + stripped[1:-1]
            if stripped:
                try:
                    return float(stripped)
                except ValueError:
                    pass
    return val


def _open_workbook(filepath: str) -> openpyxl.Workbook:
    """Load an ``.xlsx`` workbook with evaluated cell values (``data_only=True``)."""
    return openpyxl.load_workbook(filepath, data_only=True)


def _read_sheet_rows(
    wb: openpyxl.Workbook,
    sheet_index: int,
    max_rows: int = 60,
) -> List[List]:
    """Read up to *max_rows* from a worksheet.

    - Expands simple merged cells (fills every cell in the range with the
      top-left value).
    - Handles nested/grouped headers: when a merged cell spans columns in
      header rows (e.g. ``"Building Values"`` over cols 5-8), the value is
      propagated so the preview text shows the parent label for the LLM.
    - Strips currency/formatting artifacts from cell values via
      ``_extract_cell_value``.
    """
    if sheet_index >= len(wb.worksheets):
        raise IndexError(
            f"Sheet index {sheet_index} out of range "
            f"(workbook has {len(wb.worksheets)} sheets)"
        )
    ws = wb.worksheets[sheet_index]

    # Build merged-cell lookup: every (row, col) in a merged range → top-left value
    merged: dict[tuple[int, int], object] = {}
    for rng in ws.merged_cells.ranges:
        top_left_cell = ws.cell(rng.min_row, rng.min_col)
        val = _extract_cell_value(top_left_cell)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged[(r, c)] = val

    rows: list[list] = []
    for row_cells in ws.iter_rows(max_row=max_rows):
        row_vals = []
        for cell in row_cells:
            if (cell.row, cell.column) in merged:
                row_vals.append(merged[(cell.row, cell.column)])
            else:
                row_vals.append(_extract_cell_value(cell))
        rows.append(row_vals)

    max_cols = max((len(r) for r in rows), default=0)
    return [r + [None] * (max_cols - len(r)) for r in rows]


def _read_xls_rows(filepath: str, sheet_index: int, max_rows: int = 60) -> List[List]:
    """Read legacy ``.xls`` files via :mod:`xlrd`."""
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(sheet_index)
    return [
        [ws.cell_value(r, c) for c in range(ws.ncols)]
        for r in range(min(max_rows, ws.nrows))
    ]


def read_raw_rows(
    filepath: str,
    sheet_index: int = 0,
    max_rows: int = 60,
) -> List[List]:
    """Read rows from .xlsx or .xls, expanding merged cells."""
    ext = filepath.rsplit(".", 1)[-1].lower()
    if ext == "xlsx":
        wb = _open_workbook(filepath)
        return _read_sheet_rows(wb, sheet_index, max_rows)
    if ext == "xls":
        return _read_xls_rows(filepath, sheet_index, max_rows)
    raise ValueError(f"Unsupported format: .{ext}")


def rows_to_preview(rows: List[List], max_rows: int = 50) -> str:
    """Format rows as ``Row NN: cell1 | cell2 | ...`` for the LLM prompt."""
    lines: list[str] = []
    for i, row in enumerate(rows[:max_rows]):
        cells = [str(v).strip() if v is not None else "" for v in row]
        if not any(cells):
            continue
        lines.append(f"Row {i:02d}: {' | '.join(cells)}")
    return "\n".join(lines)


def get_all_sheet_previews(
    filepath: str,
    preview_rows: int = 5,
) -> List[Dict[str, Any]]:
    """Return sheet name, index, and a short text preview for every sheet."""
    ext = filepath.rsplit(".", 1)[-1].lower()
    previews: list[dict[str, Any]] = []

    if ext == "xlsx":
        wb = _open_workbook(filepath)
        for idx, ws in enumerate(wb.worksheets):
            rows = _read_sheet_rows(wb, idx, max_rows=preview_rows)
            previews.append(
                {
                    "index": idx,
                    "name": ws.title,
                    "preview": rows_to_preview(rows, max_rows=preview_rows),
                }
            )
    elif ext == "xls":
        wb = xlrd.open_workbook(filepath)
        for idx in range(wb.nsheets):
            ws = wb.sheet_by_index(idx)
            rows = [
                [ws.cell_value(r, c) for c in range(ws.ncols)]
                for r in range(min(preview_rows, ws.nrows))
            ]
            previews.append(
                {
                    "index": idx,
                    "name": getattr(ws, "name", f"Sheet{idx}"),
                    "preview": rows_to_preview(rows, max_rows=preview_rows),
                }
            )
    else:
        raise ValueError(f"Unsupported format: .{ext}")

    return previews


# ---------------------------------------------------------------------------
# Bedrock LLM
# ---------------------------------------------------------------------------


def _get_bedrock_client(
    region: Optional[str] = None,
    profile_name: Optional[str] = None,
) -> Any:
    """Build a Bedrock runtime client with retries and a configurable region or profile."""
    region = region or os.environ.get("AWS_REGION", _DEFAULT_REGION)
    cfg = BotoConfig(
        region_name=region,
        retries={"max_attempts": 3, "mode": "adaptive"},
        read_timeout=120,
    )
    session = (
        boto3.Session(profile_name=profile_name) if profile_name else boto3.Session()
    )
    return session.client("bedrock-runtime", config=cfg)


_SYSTEM_PROMPT = (
    "You are a JSON-only response bot. You MUST respond with a single valid JSON object. "
    "NEVER include markdown, code fences, backticks, explanations, comments, or any text "
    "outside the JSON. Your entire response must be parseable by json.loads(). "
    "Start your response with { and end with }. Nothing else."
)

_MAX_RETRIES = 2


def _call_bedrock(
    prompt: str,
    bedrock_cfg: Dict[str, Any],
    max_tokens: Optional[int] = None,
) -> Dict[str, Any]:
    """Send a prompt to Bedrock Converse and return text + metadata."""
    model_id = (
        bedrock_cfg.get("inference_profile_arn")
        or bedrock_cfg.get("model_id")
        or _DEFAULT_MODEL
    )
    client = _get_bedrock_client(
        region=bedrock_cfg.get("aws_region"),
        profile_name=bedrock_cfg.get("aws_profile"),
    )
    tokens = max_tokens or bedrock_cfg.get("max_tokens", 4096)

    logger.info("Calling Bedrock — model/profile: %s", _trunc(model_id))

    resp = client.converse(
        modelId=model_id,
        system=[{"text": _SYSTEM_PROMPT}],
        messages=[{"role": "user", "content": [{"text": prompt}]}],
        inferenceConfig={"maxTokens": tokens, "temperature": 0.0},
    )

    usage = resp.get("usage", {})
    logger.info(
        "Bedrock response — in: %s tok, out: %s tok, stop: %s",
        usage.get("inputTokens", "?"),
        usage.get("outputTokens", "?"),
        resp.get("stopReason", "?"),
    )

    raw = resp["output"]["message"]["content"][0]["text"].strip()
    raw = _strip_to_json(raw)

    return {
        "model_id": model_id,
        "input_tokens": usage.get("inputTokens"),
        "output_tokens": usage.get("outputTokens"),
        "stop_reason": resp.get("stopReason"),
        "raw_response": raw,
    }


def _strip_to_json(text: str) -> str:
    """Aggressively strip non-JSON wrapping from LLM output."""
    # Remove markdown fences
    text = re.sub(r"^```[a-z]*\n?", "", text)
    text = re.sub(r"\n?```$", "", text)
    text = text.strip()

    # If response has text before the JSON, find the first {
    first_brace = text.find("{")
    last_brace = text.rfind("}")
    if first_brace != -1 and last_brace != -1 and first_brace < last_brace:
        text = text[first_brace : last_brace + 1]

    return text


def _parse_llm_json(
    llm_resp: Dict[str, Any],
    bedrock_cfg: Dict[str, Any],
    context: str = "LLM response",
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """Parse JSON from LLM response, retrying once if it fails.

    Returns (parsed_dict, final_llm_response).
    """
    raw = llm_resp["raw_response"]

    # Attempt 1: parse the response
    try:
        return json.loads(raw), llm_resp
    except json.JSONDecodeError as err:
        logger.warning(
            "%s: JSON parse failed (%s), retrying with correction prompt", context, err
        )

        # Attempt 2: send the broken response back and ask for a fix
        correction_prompt = (
            f"Your previous response was not valid JSON. Here is what you returned:\n\n"
            f"{raw}\n\n"
            f"The JSON parse error was: {err}\n\n"
            f"Please return ONLY the corrected valid JSON object. "
            f"Start with {{ and end with }}. No markdown, no backticks, no explanation."
        )

        retry_resp = _call_bedrock(correction_prompt, bedrock_cfg)
        retry_raw = retry_resp["raw_response"]

        try:
            return json.loads(retry_raw), retry_resp
        except json.JSONDecodeError as err2:
            logger.error("%s: JSON parse failed on retry too (%s)", context, err2)
            logger.error("Raw response (attempt 1): %s", raw[:500])
            logger.error("Raw response (attempt 2): %s", retry_raw[:500])
            raise ValueError(
                f"LLM did not return valid JSON after {_MAX_RETRIES} attempts for {context}. "
                f"Last error: {err2}"
            ) from err2


def _trunc(s: str, n: int = 80) -> str:
    """Return ``s`` truncated from the left with an ellipsis prefix if longer than ``n``."""
    return s if len(s) <= n else f"...{s[-n:]}"


# ---------------------------------------------------------------------------
# Embeddings
# ---------------------------------------------------------------------------

_DEFAULT_EMBED_MODEL = "amazon.titan-embed-text-v2:0"
_EMBED_CACHE_FILE = "embeddings_cache.json"


def _embed_text(
    text: str,
    bedrock_cfg: Dict[str, Any],
) -> List[float]:
    """Embed a single text string using Bedrock Titan Embeddings."""
    model_id = bedrock_cfg.get("embedding_model_id", _DEFAULT_EMBED_MODEL)
    client = _get_bedrock_client(
        region=bedrock_cfg.get("aws_region"),
        profile_name=bedrock_cfg.get("aws_profile"),
    )
    body = json.dumps({"inputText": text})
    resp = client.invoke_model(modelId=model_id, body=body)
    result = json.loads(resp["body"].read())
    return result["embedding"]


def _embed_texts_batch(
    texts: List[str],
    bedrock_cfg: Dict[str, Any],
) -> List[List[float]]:
    """Embed multiple texts (individual calls, Titan doesn't support batch)."""
    return [_embed_text(t, bedrock_cfg) for t in texts]


def _cosine_similarity(a: List[float], b: List[float]) -> float:
    """Cosine similarity between two vectors."""
    dot = sum(x * y for x, y in zip(a, b))
    norm_a = sum(x * x for x in a) ** 0.5
    norm_b = sum(x * x for x in b) ** 0.5
    if norm_a == 0 or norm_b == 0:
        return 0.0
    return dot / (norm_a * norm_b)


def _load_embedding_cache(cache_path: Path) -> Dict[str, Any]:
    """Load the JSON embedding cache from disk, or return an empty dict if missing."""
    if cache_path.is_file():
        with open(cache_path, encoding="utf-8") as f:
            return json.load(f)
    return {}


def _save_embedding_cache(cache_path: Path, cache: Dict[str, Any]) -> None:
    """Persist the embedding cache to ``cache_path`` as JSON."""
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cache, f)


def _get_field_embeddings(
    schema: Dict[str, str],
    synonyms: Dict[str, List[str]],
    bedrock_cfg: Dict[str, Any],
    cache_path: Optional[Path] = None,
    force_rebuild: bool = False,
) -> Dict[str, Dict[str, Any]]:
    """Load embeddings from cache. Only rebuilds if forced or cache is missing.

    Returns {field: {"texts": [...], "vectors": [[...], ...]}}.
    """
    cache_file = cache_path or _SCRIPT_DIR / _EMBED_CACHE_FILE
    cache = _load_embedding_cache(cache_file)

    # If cache exists and no force rebuild → use it as-is
    if cache and not force_rebuild:
        logger.debug("Embedding cache loaded (%d fields)", len(cache))
        return cache

    # Build fresh embeddings
    field_texts: dict[str, list[str]] = {}
    for field, desc in schema.items():
        texts = [desc] + synonyms.get(field, [])
        field_texts[field] = texts

    logger.info("Building embedding cache for %d fields...", len(field_texts))
    total_texts = sum(len(t) for t in field_texts.values())
    logger.info("Embedding %d total texts (descriptions + synonyms)", total_texts)

    result: dict[str, dict] = {}
    for field, texts in field_texts.items():
        # Reuse cached vectors for unchanged fields
        cached = cache.get(field, {})
        if not force_rebuild and cached.get("texts") == texts:
            result[field] = cached
        else:
            vectors = _embed_texts_batch(texts, bedrock_cfg)
            result[field] = {"texts": texts, "vectors": vectors}

    _save_embedding_cache(cache_file, result)
    logger.info("Embedding cache saved to %s", cache_file)
    return result


def rebuild_embedding_cache(cfg: Dict[str, Any]) -> None:
    """Force-rebuild the embedding cache from current config synonyms."""
    _get_field_embeddings(
        cfg["fields"],
        cfg.get("synonyms", {}),
        cfg.get("bedrock", {}),
        force_rebuild=True,
    )


# ---------------------------------------------------------------------------
# Tiered column matching
# ---------------------------------------------------------------------------

_SIM_THRESHOLD = 0.80
_CANDIDATE_THRESHOLD = 0.60


def _normalize_header(text: str) -> str:
    """Normalize header text for comparison: lowercase, strip, collapse whitespace."""
    return re.sub(r"\s+", " ", str(text).strip().lower())


def _strip_year_from_header(text: str) -> str:
    """Remove year patterns like '2025', '(2024)' from header for matching."""
    return re.sub(r"\s*\(?\b(19|20)\d{2}\b\)?\s*", " ", text).strip()


# Keywords that indicate an expiring / prior-year column — must be skipped.
_EXPIRING_KEYWORDS = re.compile(
    r"\b(expir|expiring|expired|prior\s+year|previous\s+year|old\s+year|"
    r"prior\s+yr|prev\s+yr|exp\s|exp\.|renewal\s+from|outgoing|"
    r"current\s+expiring|expiry)\b",
    re.IGNORECASE,
)


def _is_expiring_column(header: str) -> bool:
    """Return True if the header text indicates an expiring / prior-year column."""
    return bool(_EXPIRING_KEYWORDS.search(header))


def match_columns_tiered(
    header_vals: List,
    schema: Dict[str, str],
    synonyms: Dict[str, List[str]],
    bedrock_cfg: Dict[str, Any],
    embedding_cfg: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[str, str], Dict[str, str], List[Tuple[int, str]]]:
    """Match Excel header columns to schema fields using synonyms then embeddings.

    Returns:
        matched: {col_index_str: field_name} — confident matches
        match_source: {col_index_str: "synonym"|"embedding(0.92)"} — how each was matched
        unmatched: [(col_index, header_text)] — columns that couldn't be matched
    """
    ecfg = embedding_cfg or {}
    sim_threshold = ecfg.get("similarity_threshold", _SIM_THRESHOLD)
    candidate_threshold = ecfg.get("candidate_threshold", _CANDIDATE_THRESHOLD)
    logger.debug(
        "Tier 2 embedding thresholds: similarity=%s candidate=%s",
        sim_threshold,
        candidate_threshold,
    )

    # Build synonym lookup: normalized_synonym → field_name
    syn_lookup: dict[str, str] = {}
    for field, syns in synonyms.items():
        for s in syns:
            syn_lookup[_normalize_header(s)] = field
        # Also add the field name itself and description
        syn_lookup[_normalize_header(field)] = field
        syn_lookup[_normalize_header(schema[field])] = field

    matched: dict[str, str] = {}
    match_source: dict[str, str] = {}
    unmatched_cols: list[tuple[int, str]] = []
    used_fields: set[str] = set()

    # --- Tier 1: Synonym matching ---
    skipped_expiring: list[tuple[int, str]] = []

    for col_idx, val in enumerate(header_vals):
        if val is None:
            continue
        raw = str(val).strip()
        if not raw:
            continue

        # Skip expiring / prior-year columns entirely
        if _is_expiring_column(raw):
            skipped_expiring.append((col_idx, raw))
            continue

        norm = _normalize_header(raw)
        norm_no_year = _normalize_header(_strip_year_from_header(raw))

        hit = syn_lookup.get(norm) or syn_lookup.get(norm_no_year)
        if hit and hit not in used_fields:
            matched[str(col_idx)] = hit
            match_source[str(col_idx)] = "synonym"
            used_fields.add(hit)
        else:
            unmatched_cols.append((col_idx, raw))

    if skipped_expiring:
        logger.info(
            "Skipped %d expiring/prior-year column(s): %s",
            len(skipped_expiring),
            [h for _, h in skipped_expiring],
        )

    tier1_count = len(matched)
    logger.info(
        "Tier 1 (synonyms): matched %d/%d columns",
        tier1_count,
        len([v for v in header_vals if v]),
    )

    if not unmatched_cols:
        return matched, match_source, []

    # --- Tier 2: Embedding matching ---
    remaining_fields = {f for f in schema if f not in used_fields}
    if not remaining_fields:
        return matched, match_source, unmatched_cols

    try:
        field_embeddings = _get_field_embeddings(schema, synonyms, bedrock_cfg)
    except Exception as exc:  # pylint: disable=broad-exception-caught
        # Bedrock embedding APIs may raise various transport or model errors.
        logger.warning("Embedding computation failed (%s), skipping Tier 2", exc)
        return matched, match_source, unmatched_cols

    # Embed unmatched headers
    unmatched_texts = [_strip_year_from_header(text) for _, text in unmatched_cols]
    try:
        header_vectors = _embed_texts_batch(unmatched_texts, bedrock_cfg)
    except Exception as exc:  # pylint: disable=broad-exception-caught
        logger.warning("Header embedding failed (%s), skipping Tier 2", exc)
        return matched, match_source, unmatched_cols

    still_unmatched: list[tuple[int, str]] = []

    for (col_idx, raw_text), header_vec in zip(unmatched_cols, header_vectors):
        best_field = None
        best_score = 0.0

        for field in remaining_fields:
            if field in used_fields:
                continue
            fe = field_embeddings.get(field, {})
            for vec in fe.get("vectors", []):
                score = _cosine_similarity(header_vec, vec)
                if score > best_score:
                    best_score = score
                    best_field = field

        if best_field and best_score >= sim_threshold and best_field not in used_fields:
            matched[str(col_idx)] = best_field
            match_source[str(col_idx)] = f"embedding({best_score:.2f})"
            used_fields.add(best_field)
            logger.info(
                "  Tier 2: col %d (%r) → %s (%.2f similarity)",
                col_idx,
                raw_text,
                best_field,
                best_score,
            )
        else:
            still_unmatched.append((col_idx, raw_text))

    tier2_count = len(matched) - tier1_count
    logger.info("Tier 2 (embeddings): matched %d more columns", tier2_count)

    return matched, match_source, still_unmatched


# ---------------------------------------------------------------------------
# Learned synonym saving
# ---------------------------------------------------------------------------


def save_learned_synonyms(
    new_mappings: Dict[str, str],
    header_vals: List,
    match_source: Dict[str, str],
    source_file: str,
    config_path: Optional[Path] = None,
) -> None:
    """Save LLM-discovered column mappings as learned synonyms for future runs.

    Only saves mappings that came from the LLM (not already matched by synonym/embedding).
    """
    learned_path = (config_path or _SCRIPT_DIR) / "learned_synonyms.yaml"

    existing: dict[str, list] = {}
    if learned_path.is_file():
        with open(learned_path, encoding="utf-8") as f:
            existing = yaml.safe_load(f) or {}

    today = date.today().isoformat()
    changed = False

    for col_idx_str, field in new_mappings.items():
        if col_idx_str in match_source:
            continue  # already matched by synonym or embedding, skip

        col_idx = int(col_idx_str)
        if col_idx >= len(header_vals):
            continue
        raw_val = header_vals[col_idx]
        if raw_val is None or str(raw_val).strip() in ("", "None", "nan"):
            continue
        header_text = str(raw_val).strip()
        header_no_year = _strip_year_from_header(header_text)

        if not header_no_year or header_no_year.lower() in ("none", "nan"):
            continue

        field_list = existing.setdefault(field, [])
        existing_names = {
            (e.get("name", e) if isinstance(e, dict) else str(e)).lower()
            for e in field_list
        }

        if header_no_year.lower() not in existing_names:
            field_list.append(
                {
                    "name": header_no_year,
                    "source": Path(source_file).name,
                    "date": today,
                }
            )
            changed = True
            logger.info(
                "Learned synonym: %r → %s (from %s)", header_no_year, field, source_file
            )

    if changed:
        with open(learned_path, "w", encoding="utf-8") as f:
            yaml.dump(existing, f, default_flow_style=False, allow_unicode=True)
        logger.info("Updated learned synonyms: %s", learned_path)


# ---------------------------------------------------------------------------
# Sheet detection
# ---------------------------------------------------------------------------


def detect_sov_sheet(
    filepath: str,
    bedrock_cfg: Dict[str, Any],
) -> Tuple[int, Optional[Dict[str, Any]]]:
    """Ask the LLM which single sheet is the best SOV data source.

    Returns (sheet_index, llm_response_dict). llm_response_dict is None for
    single-sheet workbooks (no LLM call needed).
    """
    previews = get_all_sheet_previews(filepath, preview_rows=5)
    if len(previews) == 1:
        logger.info("Single sheet workbook — using sheet 0 (%s)", previews[0]["name"])
        return 0, None

    sheet_info = "\n\n".join(
        f'### Sheet {p["index"]}: "{p["name"]}"\n{p["preview"]}' for p in previews
    )

    prompt = f"""A workbook has the following sheets. Pick the ONE best SOV data sheet.

{sheet_info}

RESPOND WITH THIS EXACT JSON STRUCTURE (no other text):
{{"sheet": <0-based index>, "reason": "<brief reason>"}}

Rules:
- Return exactly ONE sheet — the one with row-level property data (addresses, building values, TIV).
- YEAR-BASED SHEETS: If multiple sheets have similar SOV data for different years, pick the LATEST year.
- Ignore summary/pivot/cover/instruction sheets.
- YOUR ENTIRE RESPONSE MUST BE VALID JSON. No markdown. No backticks. No explanation. Start with {{ end with }}.
"""

    llm_resp = _call_bedrock(prompt, bedrock_cfg, max_tokens=100)
    result, llm_resp = _parse_llm_json(llm_resp, bedrock_cfg, context="sheet_detection")
    sheet = int(result.get("sheet", 0))
    reason = result.get("reason", "")
    logger.info(
        "LLM selected sheet %d (%s) — %s", sheet, previews[sheet]["name"], reason
    )
    return sheet, llm_resp


def resolve_sheet_indices(
    filepath: str,
    sheets_cfg: Any,
    bedrock_cfg: Dict[str, Any],
) -> Tuple[List[int], Optional[Dict[str, Any]]]:
    """Resolve configured sheet references to 0-based indices, or auto-detect.

    Returns (indices, sheet_detection_llm_response).
    """
    if not sheets_cfg:
        idx, llm_resp = detect_sov_sheet(filepath, bedrock_cfg)
        return [idx], llm_resp

    if not isinstance(sheets_cfg, list):
        sheets_cfg = [sheets_cfg]

    ext = filepath.rsplit(".", 1)[-1].lower()
    indices: list[int] = []

    for s in sheets_cfg:
        if isinstance(s, int):
            indices.append(s)
        elif isinstance(s, str):
            if ext == "xlsx":
                wb = _open_workbook(filepath)
                names = [ws.title for ws in wb.worksheets]
            else:
                xwb = xlrd.open_workbook(filepath)
                names = xwb.sheet_names()
            if s in names:
                indices.append(names.index(s))
            else:
                raise ValueError(f"Sheet '{s}' not found. Available: {names}")

    return indices or [0], None


# ---------------------------------------------------------------------------
# Analysis prompt
# ---------------------------------------------------------------------------


def _build_analysis_prompt(preview_text: str, schema: Dict[str, str]) -> str:
    schema_desc = "\n".join(f"  - {k}: {v}" for k, v in schema.items())
    return f"""You are an expert insurance data analyst specializing in Property Statement of Values (SOV) files.

I will give you a raw text preview of rows from an Excel file. Your job is to analyse it and return a JSON object.

## TARGET SCHEMA (normalized column names to map to):
{schema_desc}

## RAW EXCEL ROWS (format: "Row <index>: val1 | val2 | ..."):
{preview_text}

## YOUR TASK:
Analyse the rows above and return ONLY a valid JSON object with these keys:

{{
  "header_row": <0-based integer row index of the main header row, or null if none>,
  "data_start_row": <0-based integer row index where actual property/location data begins>,
  "data_end_row": <0-based integer row index of the LAST data row, or null if unknown>,
  "multi_year": <true if the file contains value columns for MORE THAN ONE policy year, else false>,
  "selected_year": <the LATEST year integer found (e.g. 2024), or null if not multi-year>,
  "all_years_found": [<sorted list of all year integers detected in column headers>],
  "column_mapping": {{
    "<col_index_as_string>": "<normalized_schema_key>"
  }},
  "skip_row_patterns": ["<pattern describing rows to skip, e.g. 'TOTAL', 'Subtotal', blank>"],
  "derived_columns": [
    {{
      "target": "<schema_key to produce>",
      "type": "<derivation type: sum | concat | split | regex_extract>",
      "source_cols": ["<col_index>", ...],
      "params": {{ <type-specific params> }}
    }}
  ],
  "sheet_notes": "<brief observation about the file structure>"
}}

## DERIVED COLUMN RULES — CRITICAL:
A derived column is needed when a schema field cannot be read from a single raw column.
Use "derived_columns" for these cases. Leave "column_mapping" for direct 1-to-1 mappings.
NEVER put the same field in both column_mapping and derived_columns.

### Derivation types and their params:

1. type="sum" — add two or more numeric columns to produce one value
   Use when: building_value = "Structure" col + "Contents" col + "Improvements" col
             tiv = building_value col + contents col + bi col (when no single TIV col exists)
   Example:
   {{ "target": "building_value", "type": "sum",
      "source_cols": ["9","10","11"], "params": {{}} }}

2. type="concat" — join string columns with a separator
   Use when: a single "address" field must be assembled from street + city + state + zip cols
   params: {{ "separator": ", " }}
   Example:
   {{ "target": "address", "type": "concat",
      "source_cols": ["2","3","4","5"], "params": {{"separator": ", "}} }}

3. type="split" — extract one part of a single column by splitting on a delimiter
   Use when: one column contains "Chicago, IL 60601" and you need city OR state OR zip separately
   params: {{ "delimiter": ",", "part_index": 0 }}   (0-based index of the part to keep)
   Example — extract city from "Chicago, IL 60601":
   {{ "target": "city", "type": "split",
      "source_cols": ["3"], "params": {{"delimiter": ",", "part_index": 0}} }}

4. type="regex_extract" — extract a substring using a regex capture group
   Use when: zip must be pulled from "Chicago, IL 60601" or year from "Built: 1998"
   params: {{ "pattern": "<Python regex with one capture group>" }}
   Example — extract ZIP:
   {{ "target": "zip", "type": "regex_extract",
      "source_cols": ["3"], "params": {{"pattern": "(\\d{{5}})"}} }}

### Key decision rules:
- If address is ONE column → direct mapping in column_mapping, no derivation needed
- If address is SPLIT across columns → use concat derivation for "address"
- If building_value is ONE column → direct mapping
- If building_value is spread across sub-components → use sum derivation
- If tiv is missing but individual value columns exist → derive tiv via sum
- derived_columns list may be empty []

## CRITICAL MULTI-YEAR RULES:
- SOV files often repeat value columns per year (e.g. "Bldg Value 2022 | Bldg Value 2024")
- If detected: set multi_year=true, selected_year=<highest year>
- In column_mapping, map ONLY columns for the LATEST year
- NEVER map the same schema key twice
- Non-value columns (location, address, occupancy) appear once — always include them

## CRITICAL EXPIRING / PRIOR-YEAR COLUMN RULES:
- NEVER map columns whose header contains any of these words: "Expiring", "Expiry", "Exp",
  "Prior Year", "Previous Year", "Old Year", "Outgoing", "Renewal From", "Current Expiring".
- These are old/expiring policy values and must be COMPLETELY IGNORED.
- Examples to SKIP: "Expiring Building Value", "Exp BI", "Prior Year TIV",
  "Expiring Rent", "Expiring Sum Insured", "Exp Contents", "Previous Year Limit"
- Only map CURRENT / RENEWAL / PROPOSED year columns.

## GENERAL RULES:
- column_mapping keys are 0-based column indices as strings ("0", "1", "2" ...)
- Only include columns that map to a schema key
- data_start_row must be AFTER header_row

## CRITICAL OUTPUT FORMAT:
- YOUR ENTIRE RESPONSE MUST BE A SINGLE VALID JSON OBJECT.
- Start with {{ and end with }}.
- NO markdown, NO backticks, NO code fences, NO explanation, NO text before or after the JSON.
- Every string value must be properly quoted. No trailing commas.
- The response must be directly parseable by Python json.loads().
"""


def analyse_sheet(
    preview_text: str,
    schema: Dict[str, str],
    bedrock_cfg: Dict[str, Any],
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """Send a sheet preview to Bedrock and return (parsed_analysis, llm_response).

    Fallback path — called when tiered matching can't resolve structure.
    """
    prompt = _build_analysis_prompt(preview_text, schema)
    llm_resp = _call_bedrock(prompt, bedrock_cfg)
    return _parse_llm_json(llm_resp, bedrock_cfg, context="sheet_analysis")


def analyse_sheet_with_prematched(
    preview_text: str,
    schema: Dict[str, str],
    pre_matched: Dict[str, str],
    match_source: Dict[str, str],
    unmatched: List[Tuple[int, str]],
    bedrock_cfg: Dict[str, Any],
) -> Tuple[Dict[str, Any], Optional[Dict[str, Any]]]:
    """Analyse a sheet with pre-matched columns, calling LLM only for structure + unmatched.

    If all columns are matched and no derivations are likely needed, skips the LLM
    entirely and builds the analysis dict directly.

    Returns (analysis_dict, llm_response_or_None).
    """
    schema_desc = "\n".join(f"  - {k}: {v}" for k, v in schema.items())

    # If everything matched, we still need header_row/data_start/skip_patterns from LLM
    # but with a much smaller prompt
    pre_match_desc = "\n".join(
        f"  col {idx} → {field} ({match_source.get(idx, 'matched')})"
        for idx, field in sorted(pre_matched.items(), key=lambda x: int(x[0]))
    )

    unmatched_desc = ""
    if unmatched:
        unmatched_desc = "\n\nUNMATCHED columns that need your mapping:\n" + "\n".join(
            f'  col {idx}: "{text}"' for idx, text in unmatched
        )

    prompt = f"""You are an expert insurance data analyst. I have a property SOV Excel sheet.
I have ALREADY matched most columns. I need you to:
1. Determine the structure (header_row, data_start_row, data_end_row)
2. Map any UNMATCHED columns below (if any)
3. Determine if any derived columns are needed (sum, concat, split, regex_extract)
4. Identify skip_row_patterns (TOTAL, subtotal rows etc.)

## TARGET SCHEMA:
{schema_desc}

## PRE-MATCHED COLUMNS (already resolved — do NOT change these):
{pre_match_desc}
{unmatched_desc}

## RAW EXCEL ROWS:
{preview_text}

RESPOND WITH THIS EXACT JSON STRUCTURE:
{{
  "header_row": <0-based row index of header, or null>,
  "data_start_row": <0-based row index where data begins>,
  "data_end_row": <0-based row index of last data row, or null>,
  "multi_year": <true/false>,
  "selected_year": <latest year int or null>,
  "all_years_found": [<year ints>],
  "additional_column_mapping": {{
    "<col_index>": "<schema_key>"
  }},
  "skip_row_patterns": ["TOTAL", "Subtotal"],
  "derived_columns": [],
  "sheet_notes": "<brief note>"
}}

RULES:
- additional_column_mapping should ONLY contain mappings for unmatched columns.
- Do NOT re-map pre-matched columns.
- derived_columns: only if a schema field needs computation from multiple columns.
- NEVER map columns with "Expiring", "Exp", "Prior Year", "Previous Year", "Outgoing" in the header.
  These are old/expiring values — skip them completely.
- YOUR ENTIRE RESPONSE MUST BE VALID JSON. Start with {{ end with }}. No markdown.
"""

    llm_resp = _call_bedrock(prompt, bedrock_cfg)
    result, llm_resp = _parse_llm_json(
        llm_resp, bedrock_cfg, context="sheet_analysis_prematched"
    )

    # Merge pre-matched + LLM additional mappings into final column_mapping
    final_mapping = dict(pre_matched)
    for idx_str, field in result.get("additional_column_mapping", {}).items():
        if idx_str not in final_mapping:
            final_mapping[idx_str] = field

    analysis = {
        "header_row": result.get("header_row"),
        "data_start_row": result.get("data_start_row"),
        "data_end_row": result.get("data_end_row"),
        "multi_year": result.get("multi_year", False),
        "selected_year": result.get("selected_year"),
        "all_years_found": result.get("all_years_found", []),
        "column_mapping": final_mapping,
        "skip_row_patterns": result.get("skip_row_patterns", []),
        "derived_columns": result.get("derived_columns", []),
        "sheet_notes": result.get("sheet_notes", ""),
    }

    return analysis, llm_resp


# ---------------------------------------------------------------------------
# Multi-year validation
# ---------------------------------------------------------------------------

_VALUE_KEYS = frozenset(
    {
        "building_value",
        "contents_value",
        "bi_value",
        "other_value",
        "tiv",
        "policy_limit",
        "deductible",
    }
)
_YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")


def _validate_year_columns(
    header_vals: List,
    col_mapping: Dict[str, str],
    selected_year: int,
) -> Dict[str, str]:
    """Drop value-column mappings whose header implies a different policy year."""
    cleaned: dict[str, str] = {}
    for idx_str, key in col_mapping.items():
        idx = int(idx_str)
        header = str(header_vals[idx]) if idx < len(header_vals) else ""
        years = [int(y) for y in _YEAR_RE.findall(header)]
        if key in _VALUE_KEYS and years and selected_year not in years:
            logger.info(
                "Dropped col %s (%r): year %s ≠ %s", idx, header, years, selected_year
            )
            continue
        cleaned[idx_str] = key
    return cleaned


# ---------------------------------------------------------------------------
# Derivations
# ---------------------------------------------------------------------------


def _apply_derivations(
    df_raw: pd.DataFrame,
    rules: List[Dict[str, Any]],
) -> Dict[str, pd.Series]:
    """Compute derived columns (sum, concat, split, regex_extract)."""
    derived: dict[str, pd.Series] = {}

    for rule in rules:
        target = rule.get("target", "")
        dtype = rule.get("type", "")
        src = rule.get("source_cols", [])
        params = rule.get("params", {})

        missing = [c for c in src if c not in df_raw.columns]
        if missing:
            logger.warning(
                "Derivation skipped for %r: missing cols %s", target, missing
            )
            continue

        try:
            if dtype == "sum":
                nums = [pd.to_numeric(df_raw[c], errors="coerce") for c in src]
                derived[target] = pd.concat(nums, axis=1).sum(axis=1, min_count=1)

            elif dtype == "concat":
                sep = params.get("separator", " ")
                parts = [df_raw[c].fillna("").astype(str).str.strip() for c in src]
                derived[target] = (
                    pd.DataFrame(parts)
                    .T.apply(
                        lambda row, separator=sep: separator.join(v for v in row if v),
                        axis=1,
                    )
                    .replace("", pd.NA)
                )

            elif dtype == "split":
                delim = params.get("delimiter", ",")
                part_idx = int(params.get("part_index", 0))
                series = df_raw[src[0]].astype(str).str.split(delim)
                derived[target] = series.apply(
                    lambda p, idx=part_idx: (
                        p[idx].strip()
                        if isinstance(p, list) and len(p) > idx
                        else pd.NA
                    )
                )

            elif dtype == "regex_extract":
                pattern = params.get("pattern", "")
                derived[target] = (
                    df_raw[src[0]]
                    .astype(str)
                    .str.extract(pattern, expand=False)
                    .str.strip()
                )

            else:
                logger.warning("Unknown derivation type %r for %r", dtype, target)
                continue

        except Exception as exc:  # pylint: disable=broad-exception-caught
            # Pandas/regex may raise varied errors depending on cell content.
            logger.warning("Derivation failed for %r: %s", target, exc)

    return derived


# ---------------------------------------------------------------------------
# Per-cell validation
# ---------------------------------------------------------------------------

_NA_STRINGS = frozenset(
    {"none", "nan", "n/a", "null", "-", "", "na", "tbd", "#n/a", "#ref!"}
)


def _clean_cell(val: object, ftype: str) -> object:
    """Validate and clean a single cell value based on its field type.

    Returns the cleaned value, or ``pd.NA`` if the value is invalid for the
    type.  Never drops the row — only nulls the individual cell.
    """
    if val is None:
        return pd.NA

    s = str(val).strip()
    if not s or s.lower() in _NA_STRINGS:
        return pd.NA

    if ftype == "amount":
        # Strip currency symbols, commas, spaces
        cleaned = _CURRENCY_RE.sub("", s)
        # Accounting negatives: (1234) → -1234
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return pd.NA

    if ftype == "integer":
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return pd.NA

    if ftype == "year":
        try:
            y = int(float(s))
            return y if 1900 <= y <= 2099 else pd.NA
        except (ValueError, TypeError):
            return pd.NA

    if ftype == "coordinate":
        try:
            v = float(s)
            return v if -180.0 <= v <= 180.0 else pd.NA
        except (ValueError, TypeError):
            return pd.NA

    if ftype == "boolean":
        # Normalize: "1.0" → "1", "0.0" → "0" for numeric booleans
        low = s.lower()
        try:
            low = str(int(float(low)))
        except (ValueError, TypeError):
            pass
        if low in ("yes", "y", "true", "1", "sprinklered"):
            return "Yes"
        if low in ("no", "n", "false", "0"):
            return "No"
        if low in ("partial", "p"):
            return "Partial"
        return pd.NA

    # text (default)
    return s


def _apply_field_types(
    df: pd.DataFrame,
    field_types: Dict[str, str],
) -> pd.DataFrame:
    """Apply per-cell validation to every column that has a type defined."""
    for col in df.columns:
        ftype = field_types.get(col, "text")
        df[col] = df[col].apply(lambda v, ft=ftype: _clean_cell(v, ft))
    return df


# ---------------------------------------------------------------------------
# DataFrame builder
# ---------------------------------------------------------------------------


def build_clean_dataframe(
    rows: List[List],
    analysis: Dict[str, Any],
    schema: Dict[str, str],
    field_types: Optional[Dict[str, str]] = None,
    sheet_name: Optional[str] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Apply LLM analysis to raw rows → (clean_df, source_ref_df).

    ``source_ref_df`` has the same shape as ``clean_df`` but each cell contains
    the Excel cell reference where the value was extracted from, e.g.
    ``'SOV 2025'!M4``.  For derived columns the reference shows the formula
    source cells.
    """
    data_start = analysis.get("data_start_row", 0)
    data_end = analysis.get("data_end_row")
    col_mapping = analysis.get("column_mapping", {})
    derived_rules = analysis.get("derived_columns", [])
    skip_patterns = [p.lower() for p in analysis.get("skip_row_patterns", [])]

    direct_idx = sorted(int(k) for k in col_mapping)
    derived_idx = sorted(
        int(c) for r in derived_rules for c in r.get("source_cols", [])
    )
    all_idx = sorted(set(direct_idx) | set(derived_idx))

    end = (data_end + 1) if data_end is not None else len(rows)
    data_rows = rows[data_start:end]

    sheet_prefix = f"'{sheet_name}'!" if sheet_name else ""

    records: list[dict] = []
    source_rows: list[int] = []  # original 0-based row indices that survived filtering

    for row_offset, row in enumerate(data_rows):
        max_col = max(all_idx, default=0)
        if len(row) < max_col + 1:
            row = row + [None] * (max_col + 1 - len(row))

        if all(v is None or str(v).strip() == "" for v in row):
            continue

        row_text = " ".join(str(v).lower() for v in row if v is not None)
        if any(pat in row_text for pat in skip_patterns if pat):
            continue

        records.append({str(i): row[i] for i in all_idx})
        # Excel row = data_start + row_offset + 1 (Excel is 1-based)
        source_rows.append(data_start + row_offset + 1)

    if not records:
        empty = pd.DataFrame(columns=list(schema.keys()))
        return empty, empty.copy()

    df_raw = pd.DataFrame(records)

    # Direct column mappings
    direct = {
        key: df_raw[idx] for idx, key in col_mapping.items() if idx in df_raw.columns
    }

    # Build source references for direct mappings
    direct_refs: dict[str, list[str]] = {}
    for idx_str, key in col_mapping.items():
        if idx_str in df_raw.columns:
            col_letter = _col_letter(int(idx_str))
            direct_refs[key] = [
                f"{sheet_prefix}{col_letter}{excel_row}" for excel_row in source_rows
            ]

    # Derived columns
    derived = _apply_derivations(df_raw, derived_rules) if derived_rules else {}

    # Build source references for derived columns
    derived_refs: dict[str, list[str]] = {}
    for rule in derived_rules:
        target = rule.get("target", "")
        src_cols = rule.get("source_cols", [])
        if target not in derived:
            continue
        src_letters = [_col_letter(int(c)) for c in src_cols]
        refs: list[str] = []
        for excel_row in source_rows:
            cell_refs = [f"{sheet_prefix}{ltr}{excel_row}" for ltr in src_letters]
            refs.append(", ".join(cell_refs))
        derived_refs[target] = refs

    all_cols = {**direct, **derived}
    all_refs = {**direct_refs, **derived_refs}

    # ALL schema fields as columns, in config order — missing fields get NA
    n_rows = len(df_raw)
    final_keys = list(schema.keys())

    df = pd.DataFrame({
        k: all_cols.get(k, pd.Series([pd.NA] * n_rows)) for k in final_keys
    })
    df_sources = pd.DataFrame({
        k: all_refs.get(k, [""] * n_rows) for k in final_keys
    })

    # Apply per-cell type validation
    if field_types:
        df = _apply_field_types(df, field_types)
    else:
        for col in df.select_dtypes(include="object").columns:
            df[col] = (
                df[col].astype(str).str.strip().replace({"None": pd.NA, "nan": pd.NA})
            )

    return df, df_sources


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------


def parse_sov_file(
    filepath: Union[str, Path],
    *,
    config_path: Optional[Union[str, Path]] = None,
    config_overrides: Optional[Dict[str, Any]] = None,
    verbose: bool = True,
    preview_rows: int = 60,
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, Any]]:
    """
    Parse a property SOV Excel file end-to-end.

    Returns
    -------
    tuple[pandas.DataFrame, pandas.DataFrame, dict]
        ``clean_df`` with normalized columns, ``source_ref_df`` with Excel cell
        references per cell, and ``metadata`` containing:
        ``llm_responses``, ``analyses``, and ``sheet_indices``.
    """
    configure_logging()

    fp = str(filepath)
    cfg = load_config(config_path)

    # Apply overrides
    if config_overrides:
        bedrock = cfg.setdefault("bedrock", {})
        for k, v in config_overrides.items():
            if k in (
                "inference_profile_arn",
                "aws_region",
                "aws_profile",
                "model_id",
                "max_tokens",
            ):
                bedrock[k] = v
            elif k == "sheets":
                cfg["sheets"] = v
            else:
                cfg[k] = v

    schema = cfg["fields"]
    field_types = cfg.get("field_types", {})
    synonyms = cfg.get("synonyms", {})
    bedrock_cfg = cfg.get("bedrock", {})
    embedding_cfg = cfg.get("embedding", {})

    llm_responses: list[dict] = []
    analyses: list[dict] = []

    # --- Resolve sheets ---
    sheet_indices, detect_resp = resolve_sheet_indices(
        fp, cfg.get("sheets"), bedrock_cfg
    )
    if detect_resp:
        llm_responses.append({"step": "sheet_detection", **detect_resp})
    logger.info("Target sheet(s): %s", sheet_indices)

    # --- Get sheet names for source references ---
    ext = fp.rsplit(".", 1)[-1].lower()
    sheet_names: dict[int, str] = {}
    if ext == "xlsx":
        wb = _open_workbook(fp)
        for i, ws in enumerate(wb.worksheets):
            sheet_names[i] = ws.title
    elif ext == "xls":
        xwb = xlrd.open_workbook(fp)
        for i in range(xwb.nsheets):
            sheet_names[i] = xwb.sheet_by_index(i).name

    # --- Parse each sheet ---
    all_dfs: list[pd.DataFrame] = []
    all_sources: list[pd.DataFrame] = []

    for sheet_idx in sheet_indices:
        sname = sheet_names.get(sheet_idx, f"Sheet{sheet_idx}")
        logger.info("--- Processing sheet %d (%s) ---", sheet_idx, sname)
        rows = read_raw_rows(fp, sheet_idx, max_rows=preview_rows)
        preview = rows_to_preview(rows, max_rows=preview_rows)

        # --- Tiered column matching ---
        # Heuristic: find header row (first row with >3 non-empty cells)
        header_row_idx = None
        for ri, row in enumerate(rows):
            non_empty = [v for v in row if v is not None and str(v).strip()]
            if len(non_empty) >= 3:
                header_row_idx = ri
                break
        header_vals = rows[header_row_idx] if header_row_idx is not None else []

        pre_matched, match_source, unmatched = match_columns_tiered(
            header_vals,
            schema,
            synonyms,
            bedrock_cfg,
            embedding_cfg,
        )

        # Use reduced LLM prompt with pre-matched columns
        logger.info(
            "Tiered matching: %d matched, %d unmatched → calling LLM for structure",
            len(pre_matched),
            len(unmatched),
        )
        analysis, analysis_resp = analyse_sheet_with_prematched(
            preview,
            schema,
            pre_matched,
            match_source,
            unmatched,
            bedrock_cfg,
        )
        if analysis_resp:
            llm_responses.append(
                {"step": f"sheet_{sheet_idx}_analysis", **analysis_resp}
            )
        analyses.append(analysis)

        # Save learned synonyms from LLM-discovered mappings
        config_dir = Path(config_path).parent if config_path else _SCRIPT_DIR
        save_learned_synonyms(
            analysis.get("column_mapping", {}),
            header_vals,
            match_source,
            fp,
            config_path=config_dir,
        )

        if verbose:
            _log_analysis(analysis, sheet_idx)

        # Multi-year filter
        if analysis.get("multi_year") and analysis.get("selected_year"):
            sel_year = int(analysis["selected_year"])
            hdr_idx = analysis.get("header_row")
            hdr_vals_yr = rows[hdr_idx] if hdr_idx is not None else []
            orig = len(analysis["column_mapping"])
            analysis["column_mapping"] = _validate_year_columns(
                hdr_vals_yr,
                analysis["column_mapping"],
                sel_year,
            )
            dropped = orig - len(analysis["column_mapping"])
            if dropped and verbose:
                logger.info("Removed %d stale year column(s)", dropped)

        df, df_src = build_clean_dataframe(
            rows, analysis, schema, field_types, sheet_name=sname
        )
        if not df.empty:
            all_dfs.append(df)
            all_sources.append(df_src)
        logger.info(
            "Sheet %d: %d rows, %d columns", sheet_idx, len(df), len(df.columns)
        )

    metadata = {
        "sheet_indices": sheet_indices,
        "analyses": analyses,
        "llm_responses": llm_responses,
    }

    # --- Merge ---
    if not all_dfs:
        return pd.DataFrame(), pd.DataFrame(), metadata

    if len(all_dfs) == 1:
        return all_dfs[0], all_sources[0], metadata

    # Multi-sheet merge: if shared location_id, merge on it; else concat
    if all("location_id" in df.columns for df in all_dfs):
        merged = all_dfs[0]
        merged_src = all_sources[0]
        for df, df_src in zip(all_dfs[1:], all_sources[1:]):
            new_cols = [
                c for c in df.columns if c not in merged.columns and c != "location_id"
            ]
            if new_cols:
                merged = merged.merge(
                    df[["location_id"] + new_cols],
                    on="location_id",
                    how="left",
                )
                merged_src = merged_src.merge(
                    df_src[["location_id"] + new_cols],
                    on="location_id",
                    how="left",
                )
        logger.info(
            "Merged %d sheets on location_id → %d rows", len(all_dfs), len(merged)
        )
        return merged, merged_src, metadata

    result = pd.concat(all_dfs, ignore_index=True)
    result_src = pd.concat(all_sources, ignore_index=True)
    logger.info("Concatenated %d sheets → %d rows", len(all_dfs), len(result))
    return result, result_src, metadata


def _log_analysis(analysis: Dict[str, Any], sheet_idx: int = 0) -> None:
    """Log a human-readable summary of LLM sheet analysis at INFO level."""
    logger.info("")
    logger.info("--- LLM analysis (sheet %d) ---", sheet_idx)
    logger.info("Header row       : %s", analysis.get("header_row"))
    logger.info("Data start row   : %s", analysis.get("data_start_row"))
    logger.info("Data end row     : %s", analysis.get("data_end_row", "auto"))
    if analysis.get("multi_year"):
        logger.info(
            "Multi-year       : yes, years: %s", analysis.get("all_years_found", [])
        )
        logger.info("Selected year    : %s", analysis.get("selected_year"))
    else:
        logger.info("Multi-year       : no")
    logger.info("Mapped columns   : %d", len(analysis.get("column_mapping", {})))
    for col_idx, norm in sorted(
        analysis.get("column_mapping", {}).items(),
        key=lambda x: int(x[0]),
    ):
        logger.info("  col %s → %s", col_idx, norm)
    if analysis.get("derived_columns"):
        logger.info("Derived columns  : %d", len(analysis["derived_columns"]))
        for d in analysis["derived_columns"]:
            logger.info("  %s via %s(%s)", d["target"], d["type"], d.get("source_cols"))
    if analysis.get("skip_row_patterns"):
        logger.info("Skip patterns    : %s", analysis["skip_row_patterns"])
    if analysis.get("sheet_notes"):
        logger.info("Notes            : %s", analysis["sheet_notes"])
    logger.info("--- end ---")
